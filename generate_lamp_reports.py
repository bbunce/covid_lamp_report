import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import csv, re, os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from datetime import date, timedelta, datetime
from dateutil.relativedelta import relativedelta

# add new trusts to this list if RDE starts testing for them
# see loc codes for full site names (copy and paste)
main_locations = ['Royal Devon and Exeter NHS Foundation Trust', 'Taunton and Somerset NHS Foundation Trust', 'Northern Devon Healthcare NHS Trust']

# Give user choice of generating report for all or just main sites
print("\nCOVID LAMP Report Generation\n\nMain Sites:")
for site in main_locations:
    print(site)
user_choice = input("\nGenerate LAMP report for All or just Main sites(type All or Main)?\n")
if user_choice.lower()[0] not in ['a', 'm']:
    print("Response not recognised")
    exit()

print("LOADING DATA...")

# SQL data from COVID LAMP db
lamp_file = 'data/REPORTINGTESTS.csv'

# clean raw data file and save as new file
remove_spaces_file = []
with open(lamp_file, 'r') as f:
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        remove_spaces_line = []
        for col in row:
            remove_spaces_line.append(re.sub('  +', '', col))
        remove_spaces_file.append(remove_spaces_line)

lamp_file = 'data/lamp_reporting_data.csv'
with open(lamp_file, 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerows(remove_spaces_file)

# load cleaned data file
df_all = pd.read_csv(lamp_file)

# set column data types
# remove duplicate location2 if it equals location1
df_all[['DateRequested', 'DateResulted', 'DateSampleTaken', 'DateSampleReceived']] = df_all[['DateRequested', 'DateResulted', 'DateSampleTaken', 'DateSampleReceived']].apply(pd.to_datetime, errors='coerce')
df_all['DateRequested'] = pd.to_datetime(df_all['DateRequested'], format='%Y-%m-%d')
df_all[['TestRequestID', 'UserID', 'TestStateID', 'ResultTypeID']] = df_all[['TestRequestID', 'UserID', 'TestStateID', 'ResultTypeID']].apply(pd.to_numeric, errors='coerce')
df_all['Location2Desc'] = df_all.apply(lambda x: "" if x['Location1Desc'] == x['Location2Desc'] else x['Location2Desc'], axis=1)

# smaller df for development / report only looks at previous 3 month / time saver
df_all['DateRequested'] = pd.to_datetime(df_all['DateRequested'], format='%Y-%m-%d')
df_all = df_all.loc[df_all['DateRequested'] >= '2021-11-01']

# remove samples from dataframe by barcode
remove_samples = ['RNA30627731']
df_all = df_all[~df_all['Barcode'].isin(remove_samples)]

# get list of unique locations
locations = list(df_all['Location1Desc'].unique()) + list(df_all['Location2Desc'].unique())
locations = list(set(locations))
locations.remove('')
locations.remove(np.nan)

# generate new dataframe for each data
def get_daily_stats(df_all):
    df = df_all.copy()
    df['DateRequested'] = df['DateRequested'].dt.date
    days = df['DateRequested'].unique()
    no_tests = []
    positive = []
    negative = []
    invalid = []
    expired = []
    not_received = []

    for day in days:
        no_tests.append(df.loc[(df['DateRequested'] == day) & ((df['TestStateDesc'] == 'Test complete') | (df['TestStateDesc'] == 'Sample expired'))].shape[0])
        positive.append(df.loc[(df['DateRequested'] == day) & (df['ResultDesc'] == 'Detected')].shape[0])
        negative.append(df.loc[(df['DateRequested'] == day) & (df['ResultDesc'] == 'Not Detected')].shape[0])
        invalid.append(df.loc[(df['DateRequested'] == day) & (df['ResultDesc'] == 'Invalid')].shape[0])
        expired.append(df.loc[(df['DateRequested'] == day) & (df['ResultDesc'] == 'Sample Expired')].shape[0])
        not_received.append(df.loc[(df['DateRequested'] == day) & (df['ResultDesc'] == 'Sample not received')].shape[0])

    data = {'Date' : days, 'Tests processed' : no_tests, 'Positives' : positive, 'Negatives' : negative, 'Invalid' : invalid, 'Expired' : expired, 'Sample not received': not_received}
    return pd.DataFrame.from_dict(data)

# generate indidual dataframes for each location
def get_locations_daily(df, locations):
    df_locations = {}
    for loc in locations:
        try:
            # print(df[df['Location1Desc'] == loc].head(2))
            df_locations[loc] = get_daily_stats(df[(df['Location1Desc'] == loc) | (df['Location2Desc'] == loc)])
            print(f"Processing {loc}...")
        except:
            print('Fail ', loc)
            pass
    return df_locations

sdate = date.today() - relativedelta(months=3)
edate = date.today()
# pd.date_range(sdate, edate, freq='W-MON')
def get_weekly_stats(df, sdate, edate):
    df_weekly = pd.DataFrame(index=pd.date_range(sdate, edate, freq='W-MON'))
    df_weekly = df_weekly.reset_index()
    columns = list(df.columns)
    columns.remove('Date')
    for col in columns:
        df_weekly[col] = df_weekly['index'].apply(lambda x: df[col].loc[(df['Date'] >= x) & (df['Date'] <= x + timedelta(days=6))].sum())
    df_weekly['Positive Rate'] = (df_weekly['Positives'] / (df_weekly['Positives'] + df_weekly['Negatives'])) * 100
    df_weekly['Void Rate'] = ((df_weekly['Invalid'] + df_weekly['Expired']) / (df_weekly['Positives'] + df_weekly['Negatives'] + df_weekly['Invalid'] + df_weekly['Expired'])) * 100
    return df_weekly

def get_locations_weekly(df):
    df_loc_weekly = {}
    sdate = date.today() - relativedelta(months=3)
    edate = date.today()
    for loc in df:
        df_loc_weekly[loc] = get_weekly_stats(df[loc], sdate, edate)
    return df_loc_weekly

# get last mon and sun datetimes
def get_mon_sun(last_mon):
    last_mon = re.split('-|T', str(last_mon.values[0]))
    mon = datetime(int(last_mon[0]),int(last_mon[1]),int(last_mon[2]),0,0,0)
    sun = mon + timedelta(days=6) + timedelta(seconds=86399)
    return mon, sun

def generate_graphs(df_all, df_loc_weekly, locations):
    
    for loc in locations:
        print(f"Generating {loc} graphs...")
        if loc == 'All Organisations':
            df = df_loc_weekly
            mon, sun = get_mon_sun(df['index'].iloc[-2:-1])
            users = df_all[(df_all['DateRequested'] >= mon) & (df_all['DateRequested'] <= sun)]['UserID'].value_counts()
        else:
            df = df_loc_weekly[loc]
            mon, sun = get_mon_sun(df['index'].iloc[-2:-1])
            users = df_all[((df_all['Location1Desc'] == loc) | (df_all['Location2Desc'] == loc)) & (df_all['DateRequested'] >= mon) & (df_all['DateRequested'] <= sun)]['UserID'].value_counts()
        
        # drop the most recent week (i.e. week in progress)
        df.drop(df.tail(1).index, inplace=True)
        
        fig, axs = plt.subplots(2,2, figsize=(18,18))
        fig.suptitle(f'{loc} Report', fontsize=16)
        # Test processed last 12 months
        axs[0,0].plot(df['index'],df['Tests processed'])
        axs[0,0].title.set_text('Total Tests Processed')
        axs[0,0].set_ylabel('No. Tests')
        axs[0,0].set_xticks(df['index'])
        axs[0,0].tick_params(rotation = 45)
        axs[0,0].grid()
        # Positive rate last 12 months
        axs[0,1].plot(df['index'],df['Positive Rate'])
        axs[0,1].title.set_text('Positive Rate')
        axs[0,1].set_ylabel('Positive rate (%)')
        axs[0,1].set_xticks(df['index'])
        axs[0,1].tick_params(rotation = 45)
        axs[0,1].grid()
        # if not data do not generated pie chart
        if df[['Invalid', 'Expired', 'Sample not received']].iloc[-1].sum() != 0:
            # Breakdown of void samples from last week
            axs[1,0].pie(df[['Invalid', 'Expired', 'Sample not received']].iloc[-1], labels=df[['Invalid', 'Expired', 'Sample not received']].iloc[-1].index, autopct='%1.1f%%')
            axs[1,0].title.set_text('Void Sample Last Week')
        # Number of tests per user per week for last week
        axs[1,1].hist(users, bins=[1,2,3,4,5,6,7],color='#1f77b4')
        axs[1,1].title.set_text('No. Tests Per User Last Week')
        axs[1,1].set_xlabel('No. Tests')
        axs[1,1].set_ylabel('No. Users')       
        
        fig.savefig(f'reports/images/{loc}_{date.today()}.jpg')

      

def generate_excel_report(report_name, df_all_weekly, df_loc_weekly, locations):
    loc_codes = {"All Organisations":"All","Devon Partnership NHS Trust":"DPT", "Community Pharmacies":"CommPharm", "Domiciliary care":"DomCare", "DevonPCNs":"DevPCN",
    "Somerset CCG":"SomCCG", "Yeovil NHS Foundation Trust":"YDH", "Northern Devon Healthcare NHS Trust":"NDDH", "Royal Devon and Exeter NHS Foundation Trust":"RDE",
    "Royal Cornwall Hospitals NHS Trust":"RCH", "Torbay and South Devon NHS Foundation Trust":"Torbay", "Somerset PCNs":"SomPCN", "Taunton and Somerset NHS Foundation Trust":"Taunton",
    "NHS Devon CCG":"DevCCG", "Dental contractors":"Dental", "South Western Ambulance Service NHS Foundation Trust":"SWAbulence", "DCC":"DCC",
    "Optical Contractors":"Optical", "Mount Stuart Torbay":"MtStuart", "Nuffield Exeter":"NuffExe"}

    with pd.ExcelWriter(f'reports/LAMP Inidividual Reports {date.today()}_temp.xlsx') as writer:
        df_all_weekly.rename(columns={'index':'Date', 'Positive Rate':'Positive Rate (%)', 'Void Rate':'Void Rate (%)'}, inplace=True)
        df_all_weekly.set_index('Date')
        df_all_weekly['Date'] = df_all_weekly['Date'].dt.strftime('%d/%m/%Y')
        df_all_weekly = df_all_weekly.round({'Positive Rate (%)': 2, 'Void Rate (%)': 2})
        df_all_weekly.to_excel(writer, sheet_name=f'{loc_codes["All Organisations"]}')
        for loc in locations:
            print(f"Compiling {loc} report...")
            df_loc_weekly[loc].rename(columns={'index':'Date', 'Positive Rate':'Positive Rate (%)', 'Void Rate':'Void Rate (%)'}, inplace=True)
            df_loc_weekly[loc].set_index('Date')
            df_loc_weekly[loc]['Date'] = df_loc_weekly[loc]['Date'].dt.strftime('%d/%m/%Y')
            df_loc_weekly[loc] = df_loc_weekly[loc].round({'Positive Rate (%)': 2, 'Void Rate (%)': 2})
            df_loc_weekly[loc].to_excel(writer, sheet_name=f'{loc_codes[loc]}')

    temp_lamp_report = f'reports/LAMP Inidividual Reports {date.today()}_temp.xlsx'
    wb = openpyxl.load_workbook(temp_lamp_report)

    for sheet in wb:
        # delete index column
        sheet.delete_cols(1)
        # Set column widths
        for col in range(1,10):
            if col in [1,3,4,5,6]:
                column_width = 11.0
            elif col == 2:
                column_width = 13.5
            else:
                column_width = 18
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = column_width
        # Import graphs
        for loc in loc_codes:
            if loc_codes[loc] == sheet.title:
                img = openpyxl.drawing.image.Image(f'reports/images/{loc}_{date.today()}.jpg')
                img.anchor = 'A16'
                sheet.add_image(img)
        # Conditional formatting
        greenFill = PatternFill(start_color='4dc247', end_color='4dc247', fill_type='solid')
        amberFill = PatternFill(start_color='f2c80f', end_color='f2c80f', fill_type='solid')
        redFill = PatternFill(start_color='fd625e', end_color='fd625e', fill_type='solid')
        # Positive Rate
        sheet.conditional_formatting.add('H2:H14', CellIsRule(operator='equal', formula=[0], stopIfTrue=True, fill=greenFill))
        sheet.conditional_formatting.add('H2:H14', CellIsRule(operator='between', formula=[0.001, 0.099], stopIfTrue=True, fill=amberFill))
        sheet.conditional_formatting.add('H2:H14', CellIsRule(operator='greaterThan', formula=[0.099], stopIfTrue=True, fill=redFill))
        # Void Rate
        sheet.conditional_formatting.add('I2:I14', CellIsRule(operator='equal', formula=[0], stopIfTrue=True, fill=greenFill))
        sheet.conditional_formatting.add('I2:I14', CellIsRule(operator='between', formula=[0.01, 5], stopIfTrue=True, fill=amberFill))
        sheet.conditional_formatting.add('I2:I14', CellIsRule(operator='greaterThan', formula=[5], stopIfTrue=True, fill=redFill))
    wb.save(f'reports/{date.today()} LAMP Report ({report_name}).xlsx')

    # delete temp_lamp_report file
    os.remove(temp_lamp_report)


df_loc_daily = get_locations_daily(df_all, locations)
df_all_daily = get_daily_stats(df_all)
df_loc_weekly = get_locations_weekly(df_loc_daily)
df_all_weekly = get_weekly_stats(df_all_daily, sdate, edate)

if user_choice.lower().startswith("a"):
    generate_graphs(df_all, df_loc_weekly, locations)
    generate_graphs(df_all, df_all_weekly, ['All Organisations'])
    generate_excel_report("All Sites", df_all_weekly, df_loc_weekly, locations)
    print("LAMP reports generated!")
elif user_choice.lower().startswith("m"):
    generate_graphs(df_all, df_loc_weekly, main_locations)
    generate_graphs(df_all, df_all_weekly, ['All Organisations'])
    generate_excel_report("Main Sites", df_all_weekly, df_loc_weekly, main_locations)
    print("LAMP reports generated!")

